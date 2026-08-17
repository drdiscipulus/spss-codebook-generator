from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .core import CodebookResult

if TYPE_CHECKING:
    import pandas as pd


_TABLES = ("variables", "value_labels", "missing_values", "warnings")
_WINDOWS_RESERVED_NAMES = {
    "CON",
    "PRN",
    "AUX",
    "NUL",
    *(f"COM{number}" for number in range(1, 10)),
    *(f"LPT{number}" for number in range(1, 10)),
}


@dataclass(frozen=True)
class ExportOptions:
    """User-selected export destination and output format choices."""

    output_dir: Path
    output_name: str
    export_excel: bool = True
    export_csv: bool = True
    overwrite: bool = False


@dataclass(frozen=True)
class ExportResult:
    """Paths written by one export operation."""

    written_files: list[Path]


def export_codebook(result: CodebookResult, options: ExportOptions) -> ExportResult:
    """Write codebook tables to the selected Excel and/or CSV outputs."""

    if not options.export_excel and not options.export_csv:
        raise ValueError("Select at least one export format.")

    output_dir = Path(options.output_dir)
    if output_dir.exists() and not output_dir.is_dir():
        raise NotADirectoryError(f"Output location is not a directory: {output_dir}")
    output_dir.mkdir(parents=True, exist_ok=True)
    output_name = _clean_output_name(options.output_name)
    paths = expected_output_paths(output_dir, output_name, options.export_excel, options.export_csv)

    if not options.overwrite:
        existing = [path for path in paths if path.exists()]
        if existing:
            existing_names = ", ".join(path.name for path in existing)
            raise FileExistsError(f"Output file(s) already exist: {existing_names}")

    written_files: list[Path] = []
    if options.export_excel:
        excel_path = output_dir / f"{output_name}_codebook.xlsx"
        with result_to_excel_writer(excel_path) as writer:
            for table_name in _TABLES:
                table = getattr(result, table_name)
                table.to_excel(writer, sheet_name=table_name, index=False)
                _format_worksheet(writer.book[table_name])
        written_files.append(excel_path)

    if options.export_csv:
        for suffix in _TABLES:
            table = getattr(result, suffix)
            csv_path = output_dir / f"{output_name}_{suffix}.csv"
            table.to_csv(csv_path, index=False, encoding="utf-8")
            written_files.append(csv_path)

    return ExportResult(written_files=written_files)


def expected_output_paths(
    output_dir: Path,
    output_name: str,
    export_excel: bool,
    export_csv: bool,
) -> list[Path]:
    """Return every path that may be written for overwrite checks."""

    output_name = _clean_output_name(output_name)
    paths: list[Path] = []
    if export_excel:
        paths.append(output_dir / f"{output_name}_codebook.xlsx")
    if export_csv:
        paths.extend(
            [
                output_dir / f"{output_name}_variables.csv",
                output_dir / f"{output_name}_value_labels.csv",
                output_dir / f"{output_name}_missing_values.csv",
                output_dir / f"{output_name}_warnings.csv",
            ]
        )
    return paths


def result_to_excel_writer(path: Path) -> pd.ExcelWriter:
    import pandas as pd

    return pd.ExcelWriter(path, engine="openpyxl")


def _clean_output_name(output_name: str) -> str:
    """Make a user-supplied output stem safe for common file systems."""

    cleaned = output_name.strip().rstrip(". ")
    if not cleaned:
        raise ValueError("output_name must not be empty.")
    illegal = '<>:"/\\|?*'
    for char in illegal:
        cleaned = cleaned.replace(char, "_")
    if not cleaned or cleaned in {".", ".."}:
        raise ValueError("output_name must contain at least one valid character.")
    if cleaned.split(".", maxsplit=1)[0].upper() in _WINDOWS_RESERVED_NAMES:
        cleaned = f"_{cleaned}"
    return cleaned


def _format_worksheet(worksheet) -> None:
    """Apply restrained formatting that keeps generated workbooks readable."""

    header_fill = PatternFill("solid", fgColor="17324D")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False

    for index, column_cells in enumerate(worksheet.columns, start=1):
        values = ("" if cell.value is None else str(cell.value) for cell in column_cells)
        width = min(max(max((len(value) for value in values), default=0) + 2, 10), 60)
        worksheet.column_dimensions[get_column_letter(index)].width = width
